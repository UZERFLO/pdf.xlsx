"""
Multi-Format File to Excel Converter
A desktop application that converts various file formats to Excel (.xlsx)
Supports batch conversion of multiple files
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import threading
import pandas as pd
import json
import csv
from datetime import datetime
import os

# For PDF processing
try:
    import pdfplumber
    import tabula
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# For Word documents
try:
    from docx import Document
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False

# For XML/HTML
try:
    from bs4 import BeautifulSoup
    import lxml
    XML_SUPPORT = True
except ImportError:
    XML_SUPPORT = False


class FileToExcelConverter:
    """Main converter class handling all file format conversions"""

    def __init__(self):
        self.supported_formats = {
            'csv': self.convert_csv,
            'json': self.convert_json,
            'xml': self.convert_xml,
            'html': self.convert_html,
            'txt': self.convert_txt,
            'tsv': self.convert_tsv,
        }

        if PDF_SUPPORT:
            self.supported_formats['pdf'] = self.convert_pdf
        if DOCX_SUPPORT:
            self.supported_formats['docx'] = self.convert_docx

    def convert_csv(self, file_path, output_path):
        """Convert CSV to Excel"""
        df = pd.read_csv(file_path, encoding='utf-8-sig')
        df.to_excel(output_path, index=False, engine='openpyxl')
        return True

    def convert_tsv(self, file_path, output_path):
        """Convert TSV to Excel"""
        df = pd.read_csv(file_path, sep='\t', encoding='utf-8-sig')
        df.to_excel(output_path, index=False, engine='openpyxl')
        return True

    def convert_json(self, file_path, output_path):
        """Convert JSON to Excel"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Handle different JSON structures
        if isinstance(data, list):
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            # Try to find the main data array
            for key, value in data.items():
                if isinstance(value, list):
                    df = pd.DataFrame(value)
                    break
            else:
                # If no list found, convert dict to single row
                df = pd.DataFrame([data])
        else:
            raise ValueError("Unsupported JSON structure")

        df.to_excel(output_path, index=False, engine='openpyxl')
        return True

    def convert_xml(self, file_path, output_path):
        """Convert XML to Excel"""
        if not XML_SUPPORT:
            raise ImportError("BeautifulSoup4 and lxml required for XML conversion")

        with open(file_path, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'lxml-xml')

        # Try to extract tabular data
        rows = []
        for item in soup.find_all(recursive=False):
            row_data = {}
            for child in item.find_all(recursive=False):
                row_data[child.name] = child.text
            if row_data:
                rows.append(row_data)

        if rows:
            df = pd.DataFrame(rows)
        else:
            # Fallback: create simple structure
            df = pd.DataFrame([{'Content': soup.get_text()}])

        df.to_excel(output_path, index=False, engine='openpyxl')
        return True

    def convert_html(self, file_path, output_path):
        """Convert HTML tables to Excel"""
        tables = pd.read_html(file_path)

        if not tables:
            raise ValueError("No tables found in HTML file")

        # If multiple tables, save to different sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, table in enumerate(tables):
                sheet_name = f'Table_{i+1}'
                table.to_excel(writer, sheet_name=sheet_name, index=False)

        return True

    def convert_txt(self, file_path, output_path):
        """Convert TXT to Excel (assumes tabular data)"""
        try:
            # Try comma delimiter
            df = pd.read_csv(file_path, encoding='utf-8-sig')
        except:
            try:
                # Try tab delimiter
                df = pd.read_csv(file_path, sep='\t', encoding='utf-8-sig')
            except:
                # Read as plain text
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                df = pd.DataFrame({'Content': lines})

        df.to_excel(output_path, index=False, engine='openpyxl')
        return True

    def convert_pdf(self, file_path, output_path):
        """Convert PDF to Excel - use explicit column boundaries with merged Details rows"""
        import pdfplumber
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
              
        HEADERS = ["Date", "Transaction Details", "Debits", "Credits", "Balance"]

        with pdfplumber.open(file_path) as pdf:
            first_page = pdf.pages[0]

            explicit_vertical_lines = [60, 125.0, 270, 345, 450, 555]
            #print(f"Using default boundaries: {explicit_vertical_lines}")

            table_settings = {
                "vertical_strategy": "explicit",
                "horizontal_strategy": "text",
                "explicit_vertical_lines": explicit_vertical_lines,
                "text_tolerance": 3,
                "intersection_x_tolerance": 50,
                "text_y_tolerance": 5,  # Keep this lower
            }

            all_rows = []

            for page in pdf.pages:
                page_tables = page.extract_tables(table_settings=table_settings)

                for table in page_tables:
                    if not table or len(table) == 0:
                        continue

                    header_idx = None
                    for i, row in enumerate(table):
                        if row and row[0] and str(row[0]).strip().startswith("Date"):
                            header_idx = i
                            break

                    if header_idx is None:
                        continue

                    i = header_idx + 1
                    while i < len(table):
                        row = table[i]
                        #print(row)
                        if not row or all(not cell or not str(cell).strip() for cell in row):
                            i += 1
                            continue

                        # Normalize row length to 7 columns
                        if len(row) < 5:
                            row = row + [None] * (5 - len(row))
                        elif len(row) > 5:
                            row = row[:5]

                        first_cell = str(row).strip()

                        if "Brought Forward" in first_cell:
                            i += 1
                            continue

                        if "Carried Forward" in first_cell:
                            break

                        # Determine if this is a main row (has key identifying fields)
                        has_key_fields = (row[0] and str(row[0]).strip()) or \
                                        (row[1] and str(row[1]).strip()) or \
                                        (row[4] and str(row[4]).strip()) 

                        if has_key_fields:
                            details_parts = [str(row[1]).strip() if row[1] else ""]

                            j = i + 1
                            while j < len(table):
                                next_row = table[j]
                                if len(next_row) < 5:
                                    next_row = next_row + [None] * (5 - len(next_row))

                                next_has_details = next_row[1] and str(next_row[1]).strip()
                                next_empty_elsewhere = all(
                                    not next_row[k] or not str(next_row[k]).strip()
                                    for k in [0, 2, 3, 4]
                                )

                                if next_has_details and next_empty_elsewhere:
                                    details_parts.append(str(next_row[1]).strip())
                                    j += 1
                                else:
                                    break

                            row[1] = "\n".join([p for p in details_parts if p])
                            all_rows.append(row)
                            i = j
                        else:
                            i += 1

            if not all_rows:
                raise ValueError("No tables found in PDF")

            df = pd.DataFrame(all_rows, columns=HEADERS)
            df = df.dropna(how='all')

            wb = Workbook()
            ws = wb.active
            ws.title = 'All_Data'

            for col_idx, col_name in enumerate(HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            for row_idx, row in enumerate(df.iterrows(), start=2):
                for col_idx, value in enumerate(row[1], start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            wb.save(output_path)
            return True

    def convert_docx(self, file_path, output_path):
        """Convert DOCX tables to Excel"""
        if not DOCX_SUPPORT:
            raise ImportError("python-docx required for DOCX conversion")

        doc = Document(file_path)
        tables = doc.tables

        if not tables:
            raise ValueError("No tables found in DOCX file")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, table in enumerate(tables):
                data = []
                for row in table.rows:
                    data.append([cell.text for cell in row.cells])

                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    sheet_name = f'Table_{i+1}'
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        return True

    def convert_file(self, input_path, output_dir):
        """Main conversion method"""
        input_path = Path(input_path)
        extension = input_path.suffix.lower().replace('.', '')

        if extension not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {extension}")

        # Generate output filename
        output_filename = f"{input_path.stem}_converted.xlsx"
        output_path = Path(output_dir) / output_filename

        # Perform conversion
        converter_func = self.supported_formats[extension]
        converter_func(str(input_path), str(output_path))

        return str(output_path)


class ConverterGUI:
    """GUI Application for File to Excel Converter"""

    def __init__(self, root):
        self.root = root
        self.root.title("Multi-Format File to Excel Converter")
        self.root.geometry("800x600")
        self.root.resizable(True, True)

        self.converter = FileToExcelConverter()
        self.selected_files = []
        self.output_directory = None

        self.setup_ui()

    def setup_ui(self):
        """Setup the user interface"""

        # Title
        title_frame = ttk.Frame(self.root, padding="10")
        title_frame.pack(fill=tk.X)

        title_label = ttk.Label(
            title_frame,
            text="File to Excel Converter",
            font=("Arial", 16, "bold")
        )
        title_label.pack()

        subtitle_label = ttk.Label(
            title_frame,
            text="Convert multiple files to Excel format",
            font=("Arial", 10)
        )
        subtitle_label.pack()

        # Supported formats
        formats_frame = ttk.LabelFrame(self.root, text="Supported Formats", padding="10")
        formats_frame.pack(fill=tk.X, padx=10, pady=5)

        supported = list(self.converter.supported_formats.keys())
        formats_text = "Supported: " + ", ".join([f.upper() for f in supported])
        ttk.Label(formats_frame, text=formats_text, wraplength=750).pack()

        # File selection
        file_frame = ttk.LabelFrame(self.root, text="Select Files", padding="10")
        file_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Buttons
        button_frame = ttk.Frame(file_frame)
        button_frame.pack(fill=tk.X, pady=5)

        ttk.Button(
            button_frame,
            text="Add Files",
            command=self.add_files
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Clear All",
            command=self.clear_files
        ).pack(side=tk.LEFT, padx=5)

        # File list
        list_frame = ttk.Frame(file_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.file_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            selectmode=tk.EXTENDED
        )
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)

        # Output directory
        output_frame = ttk.LabelFrame(self.root, text="Output Directory", padding="10")
        output_frame.pack(fill=tk.X, padx=10, pady=5)

        dir_frame = ttk.Frame(output_frame)
        dir_frame.pack(fill=tk.X)

        self.output_label = ttk.Label(dir_frame, text="No directory selected")
        self.output_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ttk.Button(
            dir_frame,
            text="Browse",
            command=self.select_output_directory
        ).pack(side=tk.RIGHT, padx=5)

        # Progress
        progress_frame = ttk.Frame(self.root, padding="10")
        progress_frame.pack(fill=tk.X, padx=10)

        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            length=300
        )
        self.progress_bar.pack(fill=tk.X, pady=5)

        self.status_label = ttk.Label(progress_frame, text="Ready")
        self.status_label.pack()

        # Convert button
        convert_frame = ttk.Frame(self.root, padding="10")
        convert_frame.pack(fill=tk.X)

        self.convert_button = ttk.Button(
            convert_frame,
            text="Convert All Files",
            command=self.start_conversion,
            style="Accent.TButton"
        )
        self.convert_button.pack(pady=10)

    def add_files(self):
        """Add files to conversion queue"""
        filetypes = [
            ("All Supported Files", " ".join([f"*.{ext}" for ext in self.converter.supported_formats.keys()])),
            ("CSV Files", "*.csv"),
            ("JSON Files", "*.json"),
            ("XML Files", "*.xml"),
            ("HTML Files", "*.html *.htm"),
            ("Text Files", "*.txt"),
            ("TSV Files", "*.tsv"),
            ("All Files", "*.*")
        ]

        if PDF_SUPPORT:
            filetypes.insert(1, ("PDF Files", "*.pdf"))
        if DOCX_SUPPORT:
            filetypes.insert(1, ("Word Documents", "*.docx"))

        files = filedialog.askopenfilenames(
            title="Select files to convert",
            filetypes=filetypes
        )

        for file in files:
            if file not in self.selected_files:
                self.selected_files.append(file)
                self.file_listbox.insert(tk.END, Path(file).name)

    def clear_files(self):
        """Clear all selected files"""
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)

    def select_output_directory(self):
        """Select output directory"""
        directory = filedialog.askdirectory(title="Select output directory")
        if directory:
            self.output_directory = directory
            self.output_label.config(text=directory)

    def start_conversion(self):
        """Start the conversion process"""
        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select files to convert")
            return

        if not self.output_directory:
            messagebox.showwarning("No Output", "Please select an output directory")
            return

        # Run conversion in separate thread
        thread = threading.Thread(target=self.convert_files, daemon=True)
        thread.start()

    def convert_files(self):
        """Convert all selected files"""
        total = len(self.selected_files)
        success_count = 0
        failed_files = []

        self.convert_button.config(state='disabled')
        self.progress_bar['value'] = 0
        self.progress_bar['maximum'] = total

        for i, file_path in enumerate(self.selected_files):
            filename = Path(file_path).name
            self.status_label.config(text=f"Converting: {filename}")

            try:
                output_path = self.converter.convert_file(file_path, self.output_directory)
                success_count += 1
                self.status_label.config(text=f"Converted: {filename}")
            except Exception as e:
                failed_files.append((filename, str(e)))
                self.status_label.config(text=f"Failed: {filename}")

            self.progress_bar['value'] = i + 1
            self.root.update_idletasks()

        # Show completion message
        self.convert_button.config(state='normal')

        message = f"Conversion complete!\n\nSuccessful: {success_count}/{total}"
        if failed_files:
            message += "\n\nFailed files:"
            for fname, error in failed_files[:5]:  # Show first 5 errors
                message += f"\n- {fname}: {error[:500]}"

        messagebox.showinfo("Conversion Complete", message)
        self.status_label.config(text="Ready")
        self.progress_bar['value'] = 0


def main():
    """Main application entry point"""
    root = tk.Tk()
    root.state('zoomed')
    # Set style
    style = ttk.Style()
    style.theme_use('clam')  # Use modern theme

    app = ConverterGUI(root)
    root.mainloop()
    import os
    os._exit(0)

if __name__ == "__main__":
    main()
